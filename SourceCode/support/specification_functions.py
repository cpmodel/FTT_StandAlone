# -*- coding: utf-8 -*-
"""
specification_functions.py
=========================================
Set of functions to load specifications for model solution.

Functions included:
    - load_specs
        Load specifications from the specifications workbook
    - spec_settings
        Adjust specifications given settings.ini instructions
    - check_specs
        Check that the workbook specifications are all valid
"""

# Standard library imports
import os
import copy

# Third party imports
import numpy as np
from openpyxl import load_workbook


def load_specs(name, titles_dict, fd_sectors, st_sectors, scenarios):
    """
    Load specifications from the specifications workbook.

    Parameters
    ----------
    name: str
        Name of model run, and specification file to read
    titles: dictionary of classification title lists
        Dictionary containing all title classifications
    fd_sectors: list of str
        Final demand sectors to solve endogenously (otherwise spec set to
        exogenous)
    st_sectors: list of str
        Supply and transformation sectors to solve endogenously (otherwise spec
        set to exogenous)

    Returns
    ----------
    spec_dict: dictionary of numpy arrays
        Dictionary of specification switches by sector

    Notes
    ---------
    If the parameter 'fd_sectors' is set to 'all', then all final demand
    sectors will be solved endogenously. This saves listing all modules to give
    a complete endogenous solution. Similar for 'st_sectors' for supply and
    transformation.
    """
    
    #Declare list of scenarios
    scenario_list = []
    scenario_list += [x.strip() for x in scenarios.split(',')]
    
    spec_dict = {}
    
    for scen in scenario_list:
        
    # Create file name from scenario name
        spec_file = 'specifications_{}.xlsx'.format(scen)
        
        # Check that specification workbook exists
        spec_path = os.path.join('Specifications', spec_file)
        if not os.path.isfile(spec_path):
            print('Specifications file for scenario {} not found.'.format(name))
    
        spec_wb = load_workbook(spec_path)
        sheet_names = spec_wb.sheetnames
        sheet_names.remove('Cover')
    
        # Iterate through worksheets and create NumPy arrays of switch values
        spec_dict[scen] = {}
        for sheet in sheet_names:
            active = spec_wb[sheet]
            # Identify classifications, to identify NumPy array dimensions
            row_d = titles_dict[active['B3'].value]
            col_d = titles_dict[active['B4'].value]
            # Set as default all values to exogenous
            spec_dict[scen][sheet] = np.full((len(row_d), len(col_d)), 9.0)
    
            # Check that the specification matrix is in the correct place
            if active['A6'].value != 'Switches':
                print('Format of {} is incorrect'.format(spec_file))
                print('"Switches" origin of matrix should be in cell A6')
    
            # Extract switches from workbook
            count = 0
            for value in active.iter_cols(min_row=7, max_row=7+len(row_d)-1,
                                          min_col=2, max_col=2+len(col_d)-1,
                                          values_only=True):
                spec_dict[scen][sheet][:, count] = value
                count += 1
    
        # Make adjustments to the specifications, given settings.ini instructions
        spec_dict[scen] = spec_settings(spec_dict[scen], titles_dict, fd_sectors, st_sectors,
                                  sheet_names)
    
        # Check that specification values are all valid
        spec_dict[scen] = check_specs(spec_dict[scen], sheet_names)

    # Return specifications dictionary
    return spec_dict


def spec_settings(spec_dict, titles_dict, fd_sectors, st_sectors, sheet_names):
    """
    Adjust specifications given settings.ini instructions.

    If the parameter 'fd_sectors' is set to 'all', then all final demand
    sectors will be solved endogenously. This saves listing all modules to give
    a complete endogenous solution. Similar for 'st_sectors' for supply and
    transformation.
    If only a subset of sectors are specified, then those sectors not
    included will have all switches set to exogenous.

    Parameters
    ----------
    spec_dict: dictionary of numpy arrays
        Dictionary of specification switches by sector, unadjusted
    titles_dict: dictionary of classification title lists
        Dictionary containing all title classifications
    fd_sectors: list of str
        Final demand sectors to solve endogenously (otherwise spec set to
        exogenous)
    st_sectors: list of str
        Supply and transformation sectors to solve endogenously (otherwise spec
        set to exogenous)
    sheet_names: list of str
        Sheetnames of the specifications workbook

    Returns
    ----------
    spec_dict_adj: dictionary of numpy arrays
        Specification dictionary, with adjustments for sectors not identified
        in settings.ini.
    """

    # Create a new copy of the dictionary, to adjust
    spec_dict_adj = copy.deepcopy(spec_dict)

    # Create list to populate with sectors to set to exogenous
    # This will be filled by eliminating sectors which are specified in
    # settings.ini
    exog_list = sheet_names
    exog_list.remove('economy')

    # if settings.ini set to 'all' then remove all from exog_list
    if fd_sectors == 'all'.lower():
        temp_list = titles_dict['final_energy_demand_short']
        exog_list = [e for e in exog_list if e not in temp_list]

    if st_sectors == 'all'.lower():
        temp_list = titles_dict['supply_transformation_short']
        exog_list = [e for e in exog_list if e not in temp_list]

    # Now remove any sectors from exog_list if specified in settings.ini
    sector_classes = ['final_energy_demand', 'supply_transformation']

    # Transform fd/st_sectors lists to make settings.ini case insensitive
    # E.g. if 'Passenger road transport' rather than 'tr_road_pass'
    sectors_list = []
    sectors_list += [x.strip() for x in fd_sectors.split(',')]
    sectors_list += [x.strip() for x in st_sectors.split(',')]

    for c in sector_classes:
        titles_d_lower = [x.lower() for x in titles_dict[c]]
        for sec in sectors_list:
            if sec.lower() in titles_d_lower:
                index = titles_d_lower.index(sec.lower())
                exog_list.remove(titles_dict['{}_short'.format(c)][index])

    # Set all switches to exogenous for any sectors remaining in list
    for sec in exog_list:
        spec_dict_adj[sec].fill(9)

    # Return specifications dictionary
    return spec_dict_adj


def check_specs(spec_dict, titles_dict):
    """
    Check that the workbook specifications are all valid.

    A subset of specifications are available for each sector. These are given
    in the specifications workbook, but not enforced. This function corrects
    any N/A specifications, and sets to exogenous. Also any random error values
    need to be corrected.

    Parameters
    ----------
    spec_dict: dictionary of numpy arrays
        Dictionary of specification switches by sector, unadjusted
    titles_dict: dictionary of classification title lists
        Dictionary containing all title classifications

    Returns
    ----------
    spec_dict_adj: dictionary of numpy arrays
        Specification dictionary, with adjustments for N/A specifications
    """
    # Create a new copy of the dictionary, to adjust
    spec_dict_adj = copy.deepcopy(spec_dict)

    # Replace any of the following with exogenous: non-integer, value outside
    # of range(1, 9)
    for sec in spec_dict_adj:
        temp = spec_dict_adj[sec]
        bool_arr = np.logical_or(temp % 1 != 0, temp < 1, temp > 9)
        spec_dict_adj[sec] = np.where(bool_arr, 9, temp)

    # Check for specification switches designated N/A

    return spec_dict_adj


if __name__ == "__main__":

    # This is blank. Here for testing.
    print('hello world')
