# -*- coding: utf-8 -*-
"""
=========================================
titles_functions.py
=========================================
Functions to load classification titles.

Functions included in the file:
    - load_titles
        Load model classifications and titles
"""

# Standard library imports
import os


# Third party imports
from openpyxl import load_workbook
from pathlib import Path


def load_titles():
    # Ensure we're using consistent relative paths
    dir_file = os.path.dirname(os.path.realpath(__file__))
    dir_root = Path(dir_file).parents[1] 
   
    
    """ Load model classifications and titles. """

    # Declare file name
    titles_file = 'classification_titles.xlsx'

    # Check that classification titles workbook exists
    titles_path = os.path.join(dir_root, 'Utilities', 'titles', titles_file)
    if not os.path.isfile(titles_path):
        print('Classification titles file not found.')

    titles_wb = load_workbook(titles_path)
    sheet_names = titles_wb.sheetnames
    sheet_names.remove('Cover')

    # Iterate through worksheets and add to titles dictionary
    titles_dict = {}
    for sheet in sheet_names:
        active = titles_wb[sheet]
        for column_values in active.iter_cols(min_row=1, values_only=True):
            # Assigning the full names (e.g. "1 Petrol Econ")
            if column_values[0] == 'Full name':  # First row
                titles_dict[f'{sheet}'] = column_values[1:]
            # Assigning the short names (e.g. "1")
            if column_values[0] == 'Short name': # First row
                titles_dict[f'{sheet}_short'] = column_values[1:]

    # Return titles dictionary
    return titles_dict
