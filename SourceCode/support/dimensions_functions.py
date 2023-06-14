"""
dimensions_functions.py
=========================================
Functions to load dimensions names.

Functions included:
    - load_dim
        Load model dimensions
"""

# Standard library imports
import os

# Third party imports
from openpyxl import load_workbook


def load_dims():
    """ Load model dimensions """

    # Declare file name
    dims_file = 'VariableListing.xlsx'

    # Check that classification titles workbook exists
    # Note this function is being called from 'run_NEMF.py'
    dims_path = os.path.join('utilities', 'titles', dims_file)
    if not os.path.isfile(dims_path):
        print('Dimensions name file not found.')

    dims_wb = load_workbook(dims_path)
    sn = dims_wb.sheetnames

    # Iterate through worksheets and add to titles dictionary
    dims_dict = {}
    histend = {}
    forstart = {}
    domain = {}
    for sheet in sn:
        active = dims_wb[sheet]
        for value in active.iter_rows(min_row=2, values_only=True):
            dims_dict[value[0]] = value[3:7]
            domain[value[0]] = value[7]
            histend[value[0]] = value[9]
            forstart[value[0]] = value[10]

    # Return titles dictionary
    return dims_dict, histend, domain, forstart
