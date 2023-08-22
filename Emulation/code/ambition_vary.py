# -*- coding: utf-8 -*-
"""
Created on Wed Jul 26 15:29:35 2023

Code for varying ambition between S0 and S2.

This can then be exported to country files using 
changed_input_export(compare_path, master_path) from country_inputs.py

		
MEWT		Technology subsidies, ranging between -1 (full subsidy) and 0.
MEFI		Feed-in tariffs
MEWR		Regulation in GW. When capacity is over regulation, 90% of investments into new capacity is stopped
MWKA		Exogenous capacity in GW (kickstart or phase-out policies)


@author: ib400
"""
#%% packages

import os
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
import time

os.chdir("C:/Users/ib400/OneDrive - University of Exeter/Documents/GitHub/FTT_StandAlone")
#%% regional_ambition function


def regional_ambition(regions = {'ROW': 0.2}, scenarios = ['S0','S2'], new_scen_code = None): # take out S0
    
    regions = regions
    scenarios = scenarios
    new_scen_code = new_scen_code
    
    
    # trial values
    # regions = {'US': 0.5,
    #            'CN': 0.5,
    #            'DE': 0.8,
    #            'ROW': 0.2}
    #upper_scenario = upper_scenario

    # seperate scenarios
    scenario_base = scenarios[0]
    scenario_compare = scenarios[1] 
    
    # load the comparison to establish bounds to vary within
    # this actually doesn't have to be a comparison, can just be masterfile of upper limit
    # ideally should handle variety of input files
    comparison_path = f'Emulation/data/{scenario_base}_{scenario_compare}_comparison.xlsx'
    
    # laoding wb takes a lot of time and we need to load spreadsheets anyway
    #compare_wb = load_workbook(comparison_path)
    #sheet_names = compare_wb.sheetnames
    sheet_names = ['MEWR', 'MGAM', 'MWKA', 'MEFI','MEWT'] # currently requires earliest variable first (MEWR)
    
    # seperation in variable types
    simple_vars = ['MEWR', 'MEWT', 'MEFI']
    complex_vars = ['MWKA'] ## need to add other sheets and do BCET
    background_vars = ['MGAM']
    new_sheets = []
    for sheet_name in sheet_names:
        
        #new_sheets = []
        #sheet_name = 'MEFI' # check for generalisability
        var_df = pd.read_excel(comparison_path, sheet_name=sheet_name)
        #var_df = var_df.iloc[161] # testing
        
        if sheet_name in simple_vars:
            
            for row in var_df[var_df['Scenario'] == scenario_compare].index: 
                if var_df['Country'].iloc[row] in regions.keys():
                    country = var_df['Country'].iloc[row]
                    ambition = regions[country]
                else:
                    ambition = regions['ROW']
                    
                meta = var_df.iloc[row, 0:5]
                #lower_bound = var_df.iloc[row]
                upper_bound = var_df.iloc[row]
                new_level = (upper_bound[5:] * ambition) # this is currently not Gen, neg numbers
                new_level_meta = pd.concat([meta, new_level])
                new_level_meta.loc['Scenario'] = f'{new_scen_code}'
                # need to decide where to store ambition level for meta
                new_sheets.append(new_level_meta)
                
        # This will be updated once variables with -1 off switch are restructured
        if sheet_name not in simple_vars:
            pass
    
    # maybe sort chronologically like this:
    # Concatenate Series and sort columns
    # result = pd.concat(series_list, axis=1)
    # result = result.sort_index(axis=1)
    new_sheet = pd.DataFrame(new_sheets)
    
    new_scen = {f'{new_scen_code}': new_sheet}
    
    return new_scen

#%% Example usage

S3_check = regional_ambition(regions = {'US': 0.5, 'CN': 0.5, 'ROW': 0.2}, scenarios = ['S0','S3'], new_scen_code = 'S3_check') # doesn't work




            
#%% ## Possible developments

# Need to think about MGAMs - value range is a little different
# Also background variables - BCET etc. how do we vary these? Randomly, peturbation?
# load_workbook is really slow and getting sheet names from another way, even manually, would be faster
# need to add code to save ambition levels and upper scenario

#### Argument additions:
    # variables of interest?
    





