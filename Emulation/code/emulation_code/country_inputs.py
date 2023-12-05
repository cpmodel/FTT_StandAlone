# -*- coding: utf-8 -*-
"""
Created on Fri Jul 21 11:11:45 2023

Functions for taking inputs varied from baseline and creating the relevant country sheets, then
saving to Inputs folder. Currently takes a dictionary input from ambition_vary

Compare_path - currently to comparison output but will be changed to take in input
from varied ambition function

@author: ib400
"""
import os
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
import time

# Set wd
os.chdir("C:/Users/ib400/OneDrive - University of Exeter/Documents/GitHub/FTT_StandAlone")

# Local packages
from Emulation.code.ambition_vary import regional_ambition
from Emulation.code.ambition_vary import uncertainty_generator
from Emulation.code.ambition_vary import amb_gen_sheets
from Emulation.code.ambition_vary import 



#%% Exports input sheets from dictionaries that contain ambition altered inputs
# and uncertainty analysis

def changed_input_export(scenarios, master_path = "Inputs/_MasterFiles/FTT-P/FTT-P-24x70_2021_S0.xlsx", emul_input = 'batch_1'):
    
    master_path = master_path
    emul_input = emul_input
    #tqdm()
    
    # data frame to store input values
    
    scenario_list = list(scenarios.keys())
    input_vals = pd.DataFrame(index = scenario_list)
    column_extract = scenario_list[0] 
    pol_cols = list(scenarios[column_extract][0]['ambition_levels'].keys()) # this assumes all regions same
    bgr_cols = list(scenarios[column_extract][1]['scen']['scen_inputs'].keys()) # can we delete 'scen' upstream
    
    input_vals = pd.DataFrame(columns= list(bgr_cols + pol_cols), index= scenario_list)
                              
    
    
    
    for scen in scenarios.items():
        new_scen_code = scen[0] # scenario name
        new_scen = scen[1] # scenario data
        new_scen_pol_sheets = new_scen[0]['ambition_input_sheets'] # data for exports
        new_scen_pol_values = new_scen[0]['ambition_levels'] # values to save
        new_scen_backgr_sheets = new_scen[1]['scen']['scen_input_sheets'] # data for exports
        new_scen_backgr_values = new_scen[1]['scen']['scen_inputs'] # values to save
        new_scen_backgr_sheets[0].fillna(method = 'ffill', inplace = True) # this should be moved upstream
        
        # add values to input_val for storage
        input_vals.loc[new_scen_code] = list(new_scen_backgr_values.iloc[0]) + list(new_scen_pol_values.values())
        
        # comparison_output = load_workbook(filename=compare_path) # this takes a lot of time, maybe just list, but less general
        # sheet_names = comparison_output.sheetnames
        policy_inputs = ['MEWR', 'MEWT', 'MEFI'] 
        backgr_inputs = ['BCET'] # add gammas eventually
        
        for sheet_name in policy_inputs:

            # # DEBUG
            # new_scen_code = list(new_scen.keys())[0]
            # new_scen = list(new_scen.values())[0]
            # sheet_name = 'MEWT' ##### Check generalisability
            # master_path = "Inputs/_MasterFiles/FTT-P/FTT-P-24x70_2021_S0.xlsx"
            # country = 'US'
            # END
            
            # this may be more efficient than compare scenarios process, maybe retrofit
            # imports master input file, chops of titles and puts top row as columns,
            # BE becomes technology col header but dealt with later
            master = pd.read_excel(master_path, sheet_name = sheet_name, 
                                   usecols=lambda col: col not in [1], skiprows=4) # get this into input_wrangle.py
            #master.columns = ['Technology'] + list(master.columns[1:]) # do we need to rename, can we search by index later??
            # bring outside into run file
            
            #need to generalise this to handle excels as well as Dfs
            # read in dataframe of changes in new scenario, change name of df1 for better understanding
            df1 = new_scen_pol_sheets[new_scen_pol_sheets['Sheet'] == sheet_name].reset_index(drop = True)
        
            
            countries = pd.unique(df1['Country']) # list of countries to loop through
        
            for country in countries:
                #country = 'US' ## Check generalisability
                
     
                # Country dataframe to merge in
                df_country = df1[df1['Country'] == country].reset_index(drop = True)
                # drop meta data
                df_country = df_country.drop(columns = list(df_country.columns[0:3]))
                df_country = df_country.set_index('Technology')
                
                # update master df, create it and deal with instance of BE
                if country != 'BE':
                    master_start = master[master.iloc[:, 0] == country].index[0]
                else: 
                    master_start = 0
                master_end = master_start + 35 # all rows until next country
                
                
                master_df = master[master_start:master_end].reset_index(drop = True)
                # Change title of first col to match up with current country_inputs
                master_df.iloc[0, 0] = 'Technology'
                # create order data frame
                
                master_df.columns = master_df.iloc[0]
                master_order = master_df[['Technology']]
                master_df = master_df.set_index('Technology')
                master_df = master_df[1:]
    
    
                # Update the values in the larger DataFrame using the values from the smaller DataFrame
                #master_df.update(df_country)
                master_df.update(df_country)
                updated_df = master_df.reset_index()
               
    
                # Reorder after merge
                # updated_df_reordered = master_order.merge(updated_df, on = 'Technology')
                
                updated_df.columns = [''] + list(range(2001, 2101))
                updated_df = updated_df[0:24] # remove placeholders
                
                
                # add tech numbers
                for row in updated_df.index:
                    updated_df[''].iloc[row] = str(row + 1) + ' ' + updated_df[''].iloc[row]
                    
    
                # Export to input folders
                folder_path = f'Inputs/{new_scen_code}/FTT-P'
                sheet_out = sheet_name + '_' + country
                
                # Check if already exists
                if not os.path.exists(folder_path):
                    # Create if new
                    os.makedirs(folder_path)
                    
                updated_df.to_csv(folder_path + '/' + f'{sheet_out}.csv', index = False, header = True)
                print(f'Sheet {sheet_out} saved')
                
        for sheet_name in backgr_inputs:
            for i in range(1, 72, 1):
                df_country_back = new_scen_backgr_sheets[new_scen_backgr_sheets[0] == i]
                df_country_back = df_country_back.drop(0, axis = 1)
                country = df_country_back.iloc[0,0] # get country name
                df_country_back.iloc[0,0] = '' #delete for export
                df_country_back.columns = df_country_back.iloc[0] # set columns
                df_country_back = df_country_back.iloc[1:].reset_index(drop = True)
                
                # add tech numbers
                for row in df_country_back.index:
                    df_country_back[''].iloc[row] = str(row + 1) + ' ' + df_country_back[''].iloc[row]
                    
    
                # Export to input folders
                folder_path = f'Inputs/{new_scen_code}/FTT-P'
                sheet_out = sheet_name + '_' + country
                
                # Check if already exists
                if not os.path.exists(folder_path):
                    # Create if new
                    os.makedirs(folder_path)
                    
                df_country_back.to_csv(folder_path + '/' + f'{sheet_out}.csv', index = False, header = True)
                print(f'Sheet {sheet_out} saved to {folder_path}')
                
    input_vals.to_csv(f'Emulation/data/{emul_input}.csv', index = True)
    print(f'Emulation output {emul_input} saved')
            

#%% Example usage

 #  output from ambition vary, can be generalised for different input
S3_check = regional_ambition(regions = {'US': 0.5, 'CN': 0.5, 'ROW': 0.2}, scenarios = ['S0','S3'], new_scen_code = 'S3_check')
new_scen = S3_check
# baseline inputs masterfile
master_path = "Inputs/_MasterFiles/FTT-P/FTT-P-24x70_2021_S0.xlsx"

#%% example usage

changed_input_export(scenarios)


#%% Possible improvements

#### Needs updating to handle new ambition levels rather than comparison files, ideally to do both
## Generalisability
    # Handle different forms of input
    # e.g. just a new input chosen, will create function for this
## Splice warning



        # General:
            
            # check whether compare process for chopping input is less efficient
            # We need the numbers next to the technology in the country _inputs
      
        
      
        
      
        
      
        
      
        
      
        
      
        
      
        
    