# -*- coding: utf-8 -*-
"""
Created on Wed Jul 19 11:29:12 2023

@author: ib400
"""
import os
import pandas as pd
from openpyxl import load_workbook

def compare_scenarios(scen_base, scen_compare):
    # Load the workbooks
    wb1 = load_workbook(filename=scen_base)
    wb2 = load_workbook(filename=scen_compare)

    # Get the sheet names from both workbooks
    sheet_names = wb1.sheetnames

    # Initialize an empty DataFrame to store the changed rows
    changed_rows = pd.DataFrame()

    for sheet_name in sheet_names:
        # Read the sheets from both workbooks into DataFrames
        df1 = pd.read_excel(scen_base, sheet_name=sheet_name)
        df2 = pd.read_excel(scen_compare, sheet_name=sheet_name)

        # Compare the DataFrames and find the changed rows
        diff = df1.compare(df2)

        if not diff.empty:
            # Add the changed rows to the DataFrame
            diff.insert(0, 'Sheet', sheet_name)
            changed_rows = changed_rows.append(diff, ignore_index=True)

    return changed_rows

# Example usage
scen_base = "output_workbook_S0.xlsx"
scen_compare = "output_workbook_S2.xlsx"

changed_data = compare_scenarios(scen_base, scen_compare)
print(changed_data)
