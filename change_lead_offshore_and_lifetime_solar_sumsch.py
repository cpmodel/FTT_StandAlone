# -*- coding: utf-8 -*-
"""
Created on Thu May  6 09:51:58 2021

@author: Femke
Updating lifetime and lead time for solar and offshore, respectively
"""

from celib import DB1
import openpyxl
import os





#%%
def change_times(lead_time_offshore, life_time_solar, model_to_change):

    lead_time_CSP = lead_time_offshore

    dir_excel_S0 = os.path.join(model_to_change, "In\FTTAssumptions\FTT-Power\FTT-P-24x70_2021_S0.xlsx")
    dir_excel_S1 = os.path.join(model_to_change, "In\FTTAssumptions\FTT-Power\FTT-P-24x70_2021_S1.xlsx")
    dir_excel_S2 = os.path.join(model_to_change, "In\FTTAssumptions\FTT-Power\FTT-P-24x70_2021_S2.xlsx")
    # excel_sheets_to_change = [dir_excel_S0, dir_excel_S1, dir_excel_S2]
    excel_sheets_to_change = dir_excel_S2
    #     book = load_workbook(dir_excel)    
        
    #     with pd.ExcelWriter(dir_excel, engine='openpyxl') as writer:
    #         writer.book = book
        
    #         writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    #         for i in np.arange(0, 70):
    #             onshore_generation.iloc[[i]].to_excel(writer, sheet_name="MEWG", startcol=12, startrow=21+i*36, header = False, index=False)
    #             offshore_generation.iloc[[i]].to_excel(writer, sheet_name="MEWG", startcol=12, startrow=22+i*36, header = False, index=False)
    
    for dir_excel in excel_sheets_to_change:
    
    
        srcfile = openpyxl.load_workbook(dir_excel, read_only=False) #to open the excel sheet and if it has macros
        sheetname = srcfile['BCET'] #get sheetname from the file
        for i in range(70):
            sheetname.cell(row=23+i*36,column=12).value = lead_time_offshore  # Lead time of offshore is at least 2
            sheetname.cell(row=25+i*36,column=12).value = lead_time_CSP  # Lead time of CSP is at least 2
            sheetname.cell(row=24+i*36,column=11).value = life_time_solar # Typical PV lifetimes are between 25 and 40 years.
            save_lead_times = [sheetname.cell(row=j+6+i*36, column=12).value for j in range(24)]
            save_life_times = [sheetname.cell(row=j+6+i*36, column=11).value for j in range(24)]
        sheetname = srcfile['MEWA']
        for i in range(70):
            for j in range(24):
                if j == 18:
                    sheetname.cell(row=6+j+i*36,column=21).value = 0
                else:
                    sheetname.cell(row=6+j+i*36,column=21).value = 100 / save_lead_times[j] / life_time_solar  # Lead time of offshore is at least 2
                if j == 19:
                    sheetname.cell(row=25+i*36,column=j+3).value = 0
                else:
                    sheetname.cell(row=25+i*36,column=j+3).value = 100 / lead_time_CSP / save_life_times[j]  # Lead time of offshore is at least 2
                # I'm not implementing this for offshore, as we're underestimating it. 
                # sheetname.cell(row=23+i*36,column=j+2).value = 100 / lead_time_offshore / save_life_time[j]  # Lead time of offshore is at least 2

        
        srcfile.save(dir_excel)#save it as a new file, the original file is untouched and here I am saving it as xlsm(m here denotes macros).
    
    #%%
    # countries =    ["BE", "DK", "DE", "EL", "ES", "FR", "IE", "IT", "LX", "NL", "AT",
    # 			            "PT", "FI", "SW", "UK", "CZ", "EN", "CY", "LV", "LT", "HU", "MT",
    # 			            "PL", "SI", "SK", "BG", "RO", "NO", "CH", "IS", "HR", "TR", "MK",
    # 			            "US", "JA", "CA", "AU", "NZ", "RS", "RA", "CN", "IN", "MX", "BR",
    # 			            "AR", "CO", "LA", "KR", "TW", "ID", "AS", "OP", "RW", "UE", "SD",
    # 			            "NG", "SA", "ON", "OC", "MY", "KZ", "AN", "AC", "AW", "AE", "ZA",
    # 			            "EG", "DC", "KE", "UA"]
    
    countries =     ["UK", "NL", "PL"]
    
    directly_adjust_databank = False
    db_to_change = r"C:\Users\Femke\Documents\E3ME_versions\FTT-Power_update_2021\databank\C.db1"

    if directly_adjust_databank:
        # Only do this if you're in a rush
        with DB1(db_to_change, "write") as db1:
            sorted_names = db1.index.sort_values("name")  
        
            db1.user="FN"
            for j in range (1, 62):
                old_BCET = db1.get(31070000+j*100)
                old_BCET[17, 9] = lead_time_offshore           # Change lead time of offshore to 2 years (2/3 both valid)
                old_BCET[18, 8] = life_time_solar          # Change life time solar to 30 years (like BNEF, 25-40 in literature)
                db1.scratch_if_exists(31070000+j*100)  
                # Write into the copy of databank C:
                db1.write_matrix("BCET_"+countries[j], 31070000+j*100, old_BCET, "FTT-Power cost matrix")
                
                # Change MEWA concurrently
                old_MEWA = db1.get(31060000+j*100)
                old_MEWA[17, :] = 100 / lead_time_offshore / old_BCET[:, 8]
                old_MEWA[:, 18] = 100 / old_BCET[:, 9] / life_time_solar
                db1.scratch_if_exists(31060000+j*100)  
                db1.write_matrix("MEWA_"+countries[j], 31060000+j*100, old_MEWA, "FTT-Power substitution matrix")
                
    
if __name__ == '__main__':
    model_to_change = r"C:\Users\Femke\Documents\E3ME_versions\Master"
    lead_time_offshore = 3
    life_time_solar = 30
    change_times(lead_time_offshore, life_time_solar, model_to_change)
    
